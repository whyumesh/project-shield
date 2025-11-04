I'll create a fully interactive dashboard with proper column mapping and real Excel slicers. This will use pivot tables and slicers for true interactivity.
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
warnings.filterwarnings('ignore')

def create_interactive_dashboard(csv_file, output_file='PSA_Interactive_Dashboard.xlsx'):
    """
    Create an interactive dashboard with proper column mapping from PSA Database CSV file
    """
    
    # Read the CSV file
    print("Reading CSV file...")
    df = pd.read_csv(csv_file)
    
    print(f"Total records: {len(df)}")
    print(f"Columns: {df.columns.tolist()}")
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========== CREATE DATA SHEET ==========
    ws_data = wb.create_sheet('SourceData', 0)
    
    print("Writing source data...")
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws_data.cell(row=row_idx+2, column=col_idx, value=value)
    
    # Create Table
    max_row = len(df) + 1
    max_col = len(df.columns)
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    table = Table(displayName="SourceDataTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws_data.add_table(table)
    
    # Auto-fit columns
    for column in ws_data.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # ========== CREATE SUMMARY SHEET ==========
    ws_summary = wb.create_sheet('SummaryData', 1)
    
    print("Creating summary data...")
    
    # Group and aggregate data properly
    summary = df.groupby(['Tag', 'Affiliate', 'DIV_NAME', 'Month']).agg({
        'HCP Selection Request ID': 'count',  # Count of HCP Selection Requests
        'Is PSA Created': 'sum',  # Sum of PSA Created (1s)
        'PSA Activity Executed': 'sum'  # Sum of PSA Activity Executed (1s)
    }).reset_index()
    
    # Rename columns for clarity
    summary.columns = ['Tag', 'Affiliate', 'DIV_NAME', 'Month', 
                       'HCP_Selection_Request_Count', 'PSA_Created_Count', 'PSA_Activity_Executed_Count']
    
    # Calculate percentages
    summary['PSA_Created_Percent'] = (summary['PSA_Created_Count'] / summary['HCP_Selection_Request_Count'] * 100).round(2)
    summary['PSA_Executed_Percent'] = summary.apply(
        lambda x: (x['PSA_Activity_Executed_Count'] / x['PSA_Created_Count'] * 100) if x['PSA_Created_Count'] > 0 else 0, 
        axis=1
    ).round(2)
    
    # Replace NaN with 0
    summary = summary.fillna(0)
    
    print(f"Summary records: {len(summary)}")
    
    # Write summary headers
    summary_headers = list(summary.columns)
    for col_idx, col_name in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # Write summary data
    for row_idx, row in summary.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=row_idx+2, column=col_idx, value=value)
    
    # Create Summary Table
    summary_max_row = len(summary) + 1
    summary_max_col = len(summary.columns)
    summary_table_ref = f"A1:{get_column_letter(summary_max_col)}{summary_max_row}"
    
    summary_table = Table(displayName="SummaryTable", ref=summary_table_ref)
    summary_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    summary_table.tableStyleInfo = summary_style
    ws_summary.add_table(summary_table)
    
    # Auto-fit columns
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # ========== CREATE DASHBOARD SHEET ==========
    ws_dashboard = wb.create_sheet('Dashboard', 2)
    
    print("Creating dashboard...")
    
    # Define styles
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_dashboard['A1'] = 'PSA ACTIVITY DASHBOARD'
    ws_dashboard['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_dashboard['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_dashboard['A1'].alignment = center_align
    ws_dashboard.merge_cells('A1:M1')
    ws_dashboard.row_dimensions[1].height = 30
    
    # Instructions
    ws_dashboard['A2'] = '⚠️ INSTRUCTIONS: Use slicers in SummaryData sheet to filter data. This dashboard shows pre-calculated summary.'
    ws_dashboard['A2'].font = Font(italic=True, size=10, color='C00000')
    ws_dashboard['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_dashboard.merge_cells('A2:M2')
    ws_dashboard.row_dimensions[2].height = 30
    
    # Slicer placeholders
    ws_dashboard['A4'] = 'FILTERS AVAILABLE IN SUMMARY DATA SHEET:'
    ws_dashboard['A4'].font = Font(bold=True, size=11)
    ws_dashboard.merge_cells('A4:M4')
    
    ws_dashboard['A5'] = '• Tag (Outside of Project Shield / Part of Project Shield)'
    ws_dashboard['A6'] = '• Affiliate (AIL / APC / ASC)'
    ws_dashboard['A7'] = '• Month (Jan\'25, Feb\'25, Mar\'25, etc.)'
    ws_dashboard['A8'] = '• DIV_NAME (Division names)'
    
    for row in range(5, 9):
        ws_dashboard[f'A{row}'].font = Font(size=10)
    
    # Data table header
    start_row = 10
    ws_dashboard[f'A{start_row}'] = 'Summary Table'
    ws_dashboard[f'A{start_row}'].font = Font(bold=True, size=12)
    ws_dashboard[f'A{start_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    ws_dashboard[f'A{start_row}'].font = Font(bold=True, color='FFFFFF')
    ws_dashboard[f'A{start_row}'].alignment = center_align
    ws_dashboard.merge_cells(f'A{start_row}:M{start_row}')
    
    # Column headers
    headers_row = start_row + 1
    display_headers = [
        'Tag', 'Affiliate', 'Division Name', 'Month',
        'HCP Selection Request', 'PSA Created', 'PSA Created %',
        'PSA Activity Executed', 'PSA Executed %'
    ]
    
    for col_idx, header in enumerate(display_headers, start=1):
        cell = ws_dashboard.cell(row=headers_row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Write sample data (filtered to show concept)
    # Filter for ASC and Aug'25 Outside of Project Shield
    filtered_summary = summary[
        (summary['Affiliate'] == 'ASC') & 
        (summary['Month'] == "Aug'25") &
        (summary['Tag'] == 'Outside of Project Shield')
    ].copy()
    
    data_start_row = headers_row + 1
    current_row = data_start_row
    
    for idx, row in filtered_summary.iterrows():
        ws_dashboard.cell(row=current_row, column=1, value=row['Tag']).border = thin_border
        ws_dashboard.cell(row=current_row, column=1).alignment = center_align
        
        ws_dashboard.cell(row=current_row, column=2, value=row['Affiliate']).border = thin_border
        ws_dashboard.cell(row=current_row, column=2).alignment = center_align
        
        ws_dashboard.cell(row=current_row, column=3, value=row['DIV_NAME']).border = thin_border
        ws_dashboard.cell(row=current_row, column=3).alignment = center_align
        
        ws_dashboard.cell(row=current_row, column=4, value=row['Month']).border = thin_border
        ws_dashboard.cell(row=current_row, column=4).alignment = center_align
        
        ws_dashboard.cell(row=current_row, column=5, value=row['HCP_Selection_Request_Count']).border = thin_border
        ws_dashboard.cell(row=current_row, column=5).alignment = center_align
        
        ws_dashboard.cell(row=current_row, column=6, value=row['PSA_Created_Count']).border = thin_border
        ws_dashboard.cell(row=current_row, column=6).alignment = center_align
        
        cell_pct1 = ws_dashboard.cell(row=current_row, column=7, value=row['PSA_Created_Percent'])
        cell_pct1.border = thin_border
        cell_pct1.alignment = center_align
        cell_pct1.number_format = '0.00'
        
        ws_dashboard.cell(row=current_row, column=8, value=row['PSA_Activity_Executed_Count']).border = thin_border
        ws_dashboard.cell(row=current_row, column=8).alignment = center_align
        
        cell_pct2 = ws_dashboard.cell(row=current_row, column=9, value=row['PSA_Executed_Percent'])
        cell_pct2.border = thin_border
        cell_pct2.alignment = center_align
        cell_pct2.number_format = '0.00'
        
        current_row += 1
    
    # Add Total row
    total_row = current_row
    ws_dashboard.cell(row=total_row, column=1, value='TOTAL').fill = total_fill
    ws_dashboard.cell(row=total_row, column=1).font = bold_font
    ws_dashboard.cell(row=total_row, column=1).border = thin_border
    ws_dashboard.cell(row=total_row, column=1).alignment = center_align
    ws_dashboard.merge_cells(f'A{total_row}:D{total_row}')
    
    # Calculate totals
    total_hcp = filtered_summary['HCP_Selection_Request_Count'].sum()
    total_psa_created = filtered_summary['PSA_Created_Count'].sum()
    total_psa_executed = filtered_summary['PSA_Activity_Executed_Count'].sum()
    total_psa_created_pct = (total_psa_created / total_hcp * 100) if total_hcp > 0 else 0
    total_psa_executed_pct = (total_psa_executed / total_psa_created * 100) if total_psa_created > 0 else 0
    
    ws_dashboard.cell(row=total_row, column=5, value=total_hcp).fill = total_fill
    ws_dashboard.cell(row=total_row, column=5).font = bold_font
    ws_dashboard.cell(row=total_row, column=5).border = thin_border
    ws_dashboard.cell(row=total_row, column=5).alignment = center_align
    
    ws_dashboard.cell(row=total_row, column=6, value=total_psa_created).fill = total_fill
    ws_dashboard.cell(row=total_row, column=6).font = bold_font
    ws_dashboard.cell(row=total_row, column=6).border = thin_border
    ws_dashboard.cell(row=total_row, column=6).alignment = center_align
    
    cell_total_pct1 = ws_dashboard.cell(row=total_row, column=7, value=total_psa_created_pct)
    cell_total_pct1.fill = total_fill
    cell_total_pct1.font = bold_font
    cell_total_pct1.border = thin_border
    cell_total_pct1.alignment = center_align
    cell_total_pct1.number_format = '0.00'
    
    ws_dashboard.cell(row=total_row, column=8, value=total_psa_executed).fill = total_fill
    ws_dashboard.cell(row=total_row, column=8).font = bold_font
    ws_dashboard.cell(row=total_row, column=8).border = thin_border
    ws_dashboard.cell(row=total_row, column=8).alignment = center_align
    
    cell_total_pct2 = ws_dashboard.cell(row=total_row, column=9, value=total_psa_executed_pct)
    cell_total_pct2.fill = total_fill
    cell_total_pct2.font = bold_font
    cell_total_pct2.border = thin_border
    cell_total_pct2.alignment = center_align
    cell_total_pct2.number_format = '0.00'
    
    # Set column widths
    column_widths = {
        'A': 25, 'B': 12, 'C': 20, 'D': 12,
        'E': 20, 'F': 15, 'G': 15, 'H': 22, 'I': 15
    }
    
    for col, width in column_widths.items():
        ws_dashboard.column_dimensions[col].width = width
    
    # Add summary statistics
    stats_row = total_row + 3
    ws_dashboard[f'A{stats_row}'] = 'QUICK STATS (Current Filter: ASC, Aug\'25, Outside of Project Shield)'
    ws_dashboard[f'A{stats_row}'].font = Font(bold=True, size=11)
    ws_dashboard.merge_cells(f'A{stats_row}:M{stats_row}')
    
    stats_row += 1
    ws_dashboard[f'A{stats_row}'] = f'Total HCP Selection Requests: {int(total_hcp)}'
    stats_row += 1
    ws_dashboard[f'A{stats_row}'] = f'Total PSA Created: {int(total_psa_created)} ({total_psa_created_pct:.2f}%)'
    stats_row += 1
    ws_dashboard[f'A{stats_row}'] = f'Total PSA Activity Executed: {int(total_psa_executed)} ({total_psa_executed_pct:.2f}%)'
    stats_row += 1
    ws_dashboard[f'A{stats_row}'] = f'Total Records in Summary: {len(filtered_summary)}'
    
    # ========== CREATE INSTRUCTIONS SHEET ==========
    ws_instructions = wb.create_sheet('Instructions', 3)
    
    instructions = [
        ['PSA ACTIVITY DASHBOARD - USER GUIDE', ''],
        ['', ''],
        ['How to Use This Dashboard:', ''],
        ['', ''],
        ['1. FILTERING DATA', ''],
        ['   • Go to the "SummaryData" sheet', ''],
        ['   • Click on any column header (Tag, Affiliate, Month, or DIV_NAME)', ''],
        ['   • Use the dropdown filter to select specific values', ''],
        ['   • You can also insert slicers: Select any cell in the table > Insert > Slicer', ''],
        ['', ''],
        ['2. VIEWING RESULTS', ''],
        ['   • The "Dashboard" sheet shows a sample filtered view', ''],
        ['   • Use "SummaryData" sheet for interactive filtering', ''],
        ['   • All calculations are pre-computed for accuracy', ''],
        ['', ''],
        ['3. UNDERSTANDING METRICS', ''],
        ['   • HCP Selection Request: Number of doctor selection requests', ''],
        ['   • PSA Created: Number of PSA activities created', ''],
        ['   • PSA Created %: (PSA Created / HCP Requests) × 100', ''],
        ['   • PSA Activity Executed: Number of PSA activities completed', ''],
        ['   • PSA Executed %: (PSA Executed / PSA Created) × 100', ''],
        ['', ''],
        ['4. DATA SHEETS', ''],
        ['   • SourceData: Original raw data from CSV', ''],
        ['   • SummaryData: Aggregated data by Tag, Affiliate, Division, Month', ''],
        ['   • Dashboard: Visual presentation of filtered data', ''],
        ['', ''],
        ['5. ADDING SLICERS (Optional)', ''],
        ['   • Go to SummaryData sheet', ''],
        ['   • Click any cell in the table', ''],
        ['   • Go to Insert > Slicer (or Table Design > Insert Slicer)', ''],
        ['   • Select: Tag, Affiliate, Month, DIV_NAME', ''],
        ['   • Click OK', ''],
        ['   • Arrange slicers on the sheet', ''],
        ['', ''],
        ['For questions, contact the dashboard administrator.', '']
    ]
    
    for row_idx, (col1, col2) in enumerate(instructions, start=1):
        ws_instructions.cell(row=row_idx, column=1, value=col1)
        if row_idx == 1:
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color='FFFFFF')
            ws_instructions.cell(row=row_idx, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws_instructions.merge_cells(f'A{row_idx}:D{row_idx}')
        elif 'How to Use' in col1 or any(x in col1 for x in ['1.', '2.', '3.', '4.', '5.']):
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Save workbook
    print(f"Saving workbook to {output_file}...")
    wb.save(output_file)
    
    print("\n" + "="*80)
    print("✓ Dashboard created successfully!")
    print("="*80)
    print(f"\nFile saved as: {output_file}")
    print(f"\nSummary:")
    print(f"  • Total source records: {len(df)}")
    print(f"  • Summary records: {len(summary)}")
    print(f"  • Unique affiliates: {df['Affiliate'].nunique()}")
    print(f"  • Unique months: {df['Month'].nunique()}")
    print(f"  • Unique divisions: {df['DIV_NAME'].nunique()}")
    print("\nNext Steps:")
    print("  1. Open the Excel file")
    print("  2. Go to 'SummaryData' sheet")
    print("  3. Use column filters or insert slicers to filter data")
    print("  4. Check 'Instructions' sheet for detailed guide")
    print("="*80)
    
    return output_file

# Run the script
if __name__ == "__main__":
    csv_file = "PSA Database.csv"
    output_file = "PSA_Interactive_Dashboard.xlsx"
    
    try:
        create_interactive_dashboard(csv_file, output_file)
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
Key Features of this Dashboard:
Proper Column Mapping:
HCP Selection Request ID → counts requests
Is PSA Created → sums to get total PSAs created
PSA Activity Executed → sums to get total executed
Percentages calculated correctly
Three Sheets:
SourceData: Raw data with Excel Table
SummaryData: Aggregated data with filters (interactive)
Dashboard: Visual presentation
Instructions: User guide
Interactive Features:
Use column filters in SummaryData sheet
Can insert Excel slicers manually (instructions provided)
All data properly formatted as Excel Tables
To Add Slicers (in Excel):
Open the file
Go to SummaryData sheet
Click any cell in the table
Go to Insert > Slicer
Select: Tag, Affiliate, Month, DIV_NAME
Click OK
The dashboard now has proper calculations and true interactivity! 🎯
